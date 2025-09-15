#!/usr/bin/env python3
import json
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"


def get_default_coordinates(country: str) -> Dict[str, float]:
    defaults = {
        "베트남": (14.0583, 108.2772), "라오스": (19.8563, 102.4955), "캄보디아": (12.5657, 104.9910),
        "필리핀": (12.8797, 121.7740), "몽골": (46.8625, 103.8467), "우즈베키스탄": (41.3775, 64.5853),
        "키르기스스탄": (41.2044, 74.7661), "스리랑카": (7.8731, 80.7718), "미얀마": (21.9162, 95.9560),
        "방글라데시": (23.6850, 90.3563), "태국": (15.8700, 100.9925), "인도네시아": (0.7893, 113.9213),
        "인도": (20.5937, 78.9629), "네팔": (28.3949, 84.1240), "아프가니스탄": (33.9391, 67.7100),
        "이란": (32.4279, 53.6880), "가나": (7.9465, -1.0232), "케냐": (-0.0236, 37.9062),
        "탄자니아": (-6.3690, 34.8888), "세네갈": (14.4974, -14.4524), "우간다": (1.3733, 32.2903),
        "카메룬": (7.3697, 12.3547), "모잠비크": (-18.6657, 35.5296), "에티오피아": (9.1450, 40.4897),
        "말라위": (-13.2543, 34.3015), "앙골라": (-11.2027, 17.8739), "감비아": (13.4432, -15.3101),
        "기니": (9.9456, -9.6966), "DR콩고": (-4.0383, 21.7587), "콩고민주공화국": (-4.0383, 21.7587),
        "코트디브아르": (7.5400, -5.5471), "볼리비아": (-16.2902, -63.5887), "엘살바도르": (13.7942, -88.8965),
        "페루": (-9.1900, -75.0152), "키리바시": (-3.3704, -168.7340), "한국": (35.9078, 127.7669),
        "일본": (36.2048, 138.2529), "중국": (35.8617, 104.1954), "러시아": (61.5240, 105.3188),
        "미국": (37.0902, -95.7129), "요르단": (30.5852, 36.2384), "알제리": (28.0339, 1.6596),
        "파라과이": (-23.4425, -58.4438), "말레이시아": (4.2105, 101.9758), "브루나이": (4.5353, 114.7277),
        "과테말라": (15.7835, -90.2308), "파키스탄": (30.3753, 69.3451), "아르헨티나": (-34.6037, -58.3816),
    }
    lat_lng = defaults.get(country)
    if lat_lng:
        return {"lat": lat_lng[0], "lng": lat_lng[1]}
    return {"lat": 0.0, "lng": 0.0}


def to_float(value: Any) -> float:
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(",", "")
        if not s:
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def process_oda(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    # Pick sheet containing 'ODA' or '양자계속'
    sheet = None
    for name in wb.sheetnames:
        if "ODA" in name or "양자계속" in name:
            sheet = wb[name]
            break
    if sheet is None:
        # Fallback to first sheet
        sheet = wb[wb.sheetnames[0]]

    projects = []
    # Expect header at first row; iterate rows after header
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        try:
            country = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
            x = to_float(row[2])  # X (lng)
            y = to_float(row[3])  # Y (lat)
            project_name = row[4]
            project_content = row[5]
            period = row[6]
            proj_type = row[7]
            budget = to_float(row[8])
            continent = row[9]

            if not country or not project_name:
                continue

            lat = y or 0.0
            lng = x or 0.0
            if not lat or not lng:
                coords = get_default_coordinates(str(country).strip())
                lat = coords["lat"]
                lng = coords["lng"]

            project = {
                "name": str(country).strip(),
                "lat": float(lat),
                "lng": float(lng),
                "description": str(project_name).strip(),
                "category": "ODA",
                "period": str(period).strip() if period else "",
                "budget": float(budget) if budget else 0.0,
                "continent": str(continent).strip() if continent else "",
                "type": str(proj_type).strip() if proj_type else "",
                "content": str(project_content).strip() if project_content else "",
                "number": len(projects) + 1,
                "sheet": sheet.title,
            }
            projects.append(project)
        except Exception:
            # Skip problematic rows
            continue
    return projects


def process_consulting(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    # Pick sheet containing '해외기술컨설팅' or '컨설팅'; fallback to first
    sheet = None
    for name in wb.sheetnames:
        if ("해외기술컨설팅" in name) or ("컨설팅" in name):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    projects = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        try:
            status = row[2]
            country = (row[3] or "").strip() if isinstance(row[3], str) else row[3]
            x = to_float(row[4])  # X (lng)
            y = to_float(row[5])  # Y (lat)
            english_name = row[6]
            project_name = row[7]
            project_type = row[8]
            start_date = row[9]
            end_date = row[10]
            budget = to_float(row[11])
            client = row[12]

            if not country or not project_name:
                continue

            lat = y or 0.0
            lng = x or 0.0
            if not lat or not lng:
                coords = get_default_coordinates(str(country).strip())
                lat = coords["lat"]
                lng = coords["lng"]

            project = {
                "name": str(country).strip(),
                "lat": float(lat),
                "lng": float(lng),
                "description": str(project_name).strip(),
                "category": "Consulting",
                "status": str(status).strip() if status else "",
                "budget": float(budget) if budget else 0.0,
                "client": str(client).strip() if client else "",
                "startDate": str(start_date).strip() if start_date else "",
                "endDate": str(end_date).strip() if end_date else "",
                "projectType": str(project_type).strip() if project_type else "",
                "englishName": str(english_name).strip() if english_name else "",
                "number": len(projects) + 1,
                "sheet": sheet.title,
            }
            projects.append(project)
        except Exception:
            continue
    return projects


def main() -> None:
    oda_xlsx = DATA_DIR / "global ODA.xlsx"
    consulting_xlsx = DATA_DIR / "global consulting.xlsx"

    outputs = []

    if oda_xlsx.exists():
        oda_projects = process_oda(oda_xlsx)
        oda_out = DATA_DIR / "global_oda.json"
        oda_out.write_text(json.dumps(oda_projects, ensure_ascii=False, indent=2), encoding="utf-8")
        outputs.append(str(oda_out.relative_to(BASE_DIR)))

    if consulting_xlsx.exists():
        consulting_projects = process_consulting(consulting_xlsx)
        consulting_out = DATA_DIR / "global_consulting.json"
        consulting_out.write_text(json.dumps(consulting_projects, ensure_ascii=False, indent=2), encoding="utf-8")
        outputs.append(str(consulting_out.relative_to(BASE_DIR)))

    print("Generated:")
    for p in outputs:
        print(f" - {p}")


if __name__ == "__main__":
    main()

